import tkinter as tk
import re
import openpyxl
from tkinter import END, messagebox
from tkinter import ttk
import pandas as pd
from tkinter.filedialog import askopenfile
from tkcalendar import DateEntry
from openpyxl.styles import numbers
import datetime
import os
from tkinter import *


dt = datetime.date.today()
num = 0
num1 = 0
num3 = 0
result = 0
result1 = 0
pymt = 0
ff = 0
jj = 0
p = 0
pp = 0
psy = 0
dd = 0
pt = 0
lh = 0
nd = ""
pyt = 0
pya = 0



def popup():
    messagebox.showinfo("info", "Please insert the correct Title of Books")


def new_window():
    sp1 = openpyxl.load_workbook('n1.xlsx')
    ofs = sp1['Sheet1']
    ofs3 = sp1['Sheet2']
    global result1
    my_w.iconify()
    discount = tk.StringVar()
    discount1 = tk.StringVar()
    ofs.cell(3, 3, value=name.get().upper())
    ofs.cell(4, 3, value=address.get().upper())
    ofs.cell(2, 4, value=f"As of  {dt}")
    AS.cell(2, 3, value=f"As of  {dt}")
    window2 = tk.Toplevel(my_w)
    window2.geometry("675x510")
    window2.title(f"Order Form {name.get()}")

    #treeview
    tv = ttk.Treeview(window2, selectmode='browse')
    tv.grid(row=7, column=1, columnspan=5, pady=10, padx=10)
    tv["columns"] = ("1", "2", "3", "4", "5", "6", "7")
    tv['show'] = 'headings'
    tv.column("1", width=50, anchor='c')
    tv.column("2", width=50, anchor='c')
    tv.column("3", width=50, anchor='c')
    tv.column("4", width=350, anchor='c')
    tv.column("5", width=50, anchor='c')
    tv.column("6", width=50, anchor='c')
    tv.column("7", width=50, anchor='c')
    tv.heading("1", text="DATE")
    tv.heading("2", text="DR/CM")
    tv.heading("3", text="QTY")
    tv.heading("4", text="Title of Books")
    tv.heading("5", text="Unit Price")
    tv.heading("6", text="Debit")
    tv.heading("7", text="Credit")
    e1_str = tk.StringVar()
    date = DateEntry(window2, selectmode="day")
    date.grid(row=1, column=4)
    label3 = tk.Label(window2, text="Tile of Books")
    label3.grid(row=0, column=2, columnspan=1)
    labeltotal = tk.Label(window2, text="Sub. Total :")
    labeltotal1 = tk.Label(window2, text="Less Discount :")
    labeltotal2 = tk.Label(window2, text="Total :")
    labeltotal2.grid(row=10, column=3, sticky="e")
    labeltotal1.grid(row=9, column=3, sticky="e")
    labeltotal.grid(row=8, column=3, sticky="e")
    total_discount = tk.Label(window2, fg="red", text=result1)
    total_discount.grid(row=9, column=4, sticky="e")
    total_total = tk.Label(window2, text=result1)
    total_total.grid(row=10, column=4, sticky="e")
    total_label = tk.Label(window2, text=result1)
    total_label.grid(row=8, column=4, sticky="e")
    label4 = tk.Label(window2, text="QTY")
    label4.grid(row=0, column=1)
    label5 = tk.Label(window2, text="DR#/CM#")
    label5.grid(row=0, column=3)
    label6 = tk.Label(window2, text="DATE")
    label6.grid(row=0, column=4)
    dr = tk.Entry(window2, width=6)

    dr.grid(row=1, column=3)
    e1 = tk.Entry(window2, textvariable=e1_str, width=55)
    e1.grid(row=1, column=2, columnspan=1)
    e2 = tk.Entry(window2, width=5)
    e2.grid(row=1, column=1)
    l1 = tk.Listbox(window2, height=6, relief="flat", bg="SystemButtonFace", width=55)
    l1.grid(row=2, column=2, columnspan=1, rowspan=4)
    dc = tk.Entry(window2, width=5, textvariable=discount)
    dc2 = tk.Entry(window2, width=5, textvariable=discount1)
    label8 = tk.Label(window2, text="    +")
    label8.grid(row=4, column=4,columnspan=5)
    dc.insert(0, "15")
    dc2.insert(0, "0")
    dc.grid(row=4, column=4)
    dc2.grid(row=4, column=5)
    label7 = tk.Label(window2, text="Discount :")
    label7.grid(row=4, column=3)

    def data_collect(self):
        e1.delete(0, END)
        e2.delete(0, END)
        dr.delete(0, END)
        date.delete(0, END)
        selected = tv.focus()
        values = tv.item(selected, 'values')
        e1.insert(0, values[3])
        e2.insert(0, values[2])
        dr.insert(0, values[1])
        date.insert(0, values[0])

    def up_button():
        global result1
        selected = tv.focus()
        my = e1.get()
        quantity = e2.get()
        values = tv.item(selected, 'values')
        result1 -= int(values[5])
        y = 0
        if quantity == "":
            messagebox.showinfo("info", "Insert Quantity")

        for row in range(2, sheet.max_row + 1):
            cell = sheet.cell(row, 1)
            if my == cell.value:
                price = sheet.cell(row, 2)
                qty1 = int(quantity)
                total = qty1 * price.value
                tv.item(selected, text="", values=(date.get(), dr.get(), quantity, cell.value, price.value, total))
                result1 += total
                total_label.config(text="₱{:,.2f}".format(result1))
                dis = int(dc.get())
                dis2 = int(dc2.get())
                dis1 = dis / 100 * result1
                res = result1 - dis1
                dis3 = dis2 / 100 * res
                total_discount.config(text="₱{:,.2f}".format(dis1+dis3))
                res = res - dis3
                total_total.config(text="₱{:,.2f}".format(res))
            else:
                y += 1
                if y == sheet.max_row - 1:
                    popup()

    def my_upd(my_widget):
        my = my_widget.widget
        index = int(my.curselection()[0])
        value = my.get(index)
        e1_str.set(value)
        l1.delete(0, END)

    def rem_button():
        global result1
        selected = tv.focus()
        values = tv.item(selected, 'values')
        result1 -= int(values[5])
        x = tv.selection()[0]
        tv.delete(x)
        total_label.config(text="₱{:,.2f}".format(result1))
        dis = int(dc.get())
        dis2 = int(dc2.get())
        dis1 = dis / 100 * result1
        res = result1 - dis1
        dis3 = dis2 / 100 * res
        total_discount.config(text="₱{:,.2f}".format(dis1 + dis3))
        res = res - dis3
        total_total.config(text="₱{:,.2f}".format(res))

    def get_data(*args):
        search_str = e1.get()
        l1.delete(0, END)
        if search_str == "":
            l1.delete(0, END)

        else:
            for element in my_list:
                if re.match(search_str, element, re.IGNORECASE):
                    l1.insert(tk.END, element)

    def add_button(*args):
        global num, result1
        my = e1.get()
        quantity = e2.get()
        y = 0
        dr1 = dr.get()
        dt1 = date.get()

        if dr1 == "":
            dr1 = "-"

        if dt1 == "":
            dt1 = "-"

        if quantity == "":
            messagebox.showinfo("info", "Insert Quantity")

        for row in range(2, sheet.max_row + 1):

            global result1
            cell = sheet.cell(row, 1)
            if my == cell.value:
                price = sheet.cell(row, 2)
                qty1 = int(quantity)
                total = qty1 * price.value
                tv.insert("", 'end', iid=num, values=(dt1, dr1, quantity, cell.value, price.value, total, ""))
                e1.delete(0, END)
                date.delete(0, END)
                dr.delete(0, END)
                num = num + 1
                result1 += total + 0.0
                total_label.config(text="₱{:,.2f}".format(result1))
                dis = int(dc.get())
                dis2 = int(dc2.get())
                dis1 = dis / 100 * result1
                res = result1 - dis1
                dis3 = dis2 / 100 * res
                total_discount.config(text="₱{:,.2f}".format(dis1 + dis3))
                res = res - dis3
                total_total.config(text="₱{:,.2f}".format(res))

            else:
                y += 1
                if y == sheet.max_row - 1:
                    popup()

    def sub_button():
        global result, result1, p
        dis = int(dc.get())
        dis2 = int(dc2.get())
        for row in tv.get_children():
            values = tv.item(row, 'values')
            ofs.append([values[0], values[1], values[2], values[3]])
            ofs.cell(ofs.max_row, 5, value=int(values[4])).number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED2
            ofs.cell(ofs.max_row, 6, value=int(values[5])).number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED2
            ofs.cell(6, 7, value="")
            nc = ofs.cell(int(ofs.max_row), 6)
            result = result + nc.value
            ofs.cell(ofs.max_row, 8, value=result).number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED2



        dis1 = dis / 100 * result
        res = result - dis1
        dis3 = dis2 / 100 * res
        ofs.append(["----", "", "------", f"LESS {dc.get()}% DISSCOUNT", "-----", "", "₱ - {:,.2f}".format(dis1),
                    "            {:,.2f}".format(res)])
        if dis2 == 0:
            pass
        else:
            res = res - dis3
            ofs.append(["----", "", "------", f"LESS {dc2.get()}% DISSCOUNT", "-----", "", "₱ - {:,.2f}".format(dis1),
                    "            {:,.2f}".format(res)])
        ofs.append(["----", "---", "---", "TOTAL BALANCE", "-----", "---", " --- ", "₱         {:,.2f}".format(res)])
        ofs.append([""])
        ofs.append([""])
        ofs.append(["", "Received by:", "", "", "", "", "Prepared by:"])
        ofs.append([""])
        ofs.append(["", "____________", "", "", "", "", "_____________"])
        AS.delete_rows(idx=AS.max_row)
        for row in range(4, AS.max_row + 1):
            s = name.get().upper()
            ad = address.get().upper()
            a = row + 1
            cell = AS.cell(row, 2)

            if s == cell.value:
                AS.cell(a, 4, value=res).number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED2

            if row == AS.max_row:
                AS.cell(a, 2, value=s)
                AS.cell(a, 3, value=ad)
                AS.cell(a, 4, value=res).number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED2

        for row in range(4, AS.max_row):
            x = row - 3
            cell = AS.cell(row + 1, 4)
            p = p + int(cell.value)
            AS.cell(row + 1, 1, value=x)
            if row == AS.max_row-1:
                AS.cell(row+2, 3, value="TOTAL :")
                AS.cell(row+2, 4, value=p).number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED2
        ofs3.cell(1, 1, value=dis)
        ofs3.cell(1, 2, value=dis2)
        sp2.save("sp.xlsx")
        sp1.save(f'{name.get().upper()}.xlsx')
        messagebox.showinfo("info", "File Saved!!")

        my_w.destroy()

    #Buttons
    b2 = tk.Button(window2, text=" Submit ", command=sub_button)
    b1 = tk.Button(window2, text="  Add  ", command=add_button)
    up = tk.Button(window2, text=" Update ", command=up_button)
    rem = tk.Button(window2, text="Remove", command=rem_button)
    rem.grid(row=8, column=1, columnspan=1)
    up.grid(row=6, column=3, columnspan=1)
    b2.grid(row=8, column=2, columnspan=2)
    b1.grid(row=6, column=2, columnspan=1)

    l1.bind("<<ListboxSelect>>", my_upd)
    e1_str.trace('w', get_data)
    tv.bind("<<TreeviewSelect>>", data_collect)
    window2.bind("<Return>",add_button )

    def on_close():

        close = messagebox.askokcancel("Close", "Would you like to close the program?")
        if close:
            my_w.destroy()

    window2.protocol("WM_DELETE_WINDOW", on_close)


sp2 = openpyxl.load_workbook('sp.xlsx')
AS = sp2['Sheet1']

sp = openpyxl.load_workbook('sapi.xlsx')
sheet = sp['Sheet1']


filename = r'sapi.xlsx'
dfe = pd.read_excel(filename, sheet_name="Sheet1")


my_w = tk.Tk()
my_w.geometry("500x200")
my_w.title("Sapi")
font = ('Times', 24, 'bold')
my_list = list(dfe['Title'])
label = tk.Label(my_w, text="Name of Your School :")
label2 = tk.Label(my_w, text="")
label.grid(row=1, column=1)
label2.grid(row=0, column=1)
name = tk.Entry(my_w, width=50)
name.grid(row=1, column=2)
label1 = tk.Label(my_w, text="Address :")


label1.grid(row=2, column=1, sticky="e",pady=5)
address = tk.Entry(my_w, width=50)
address.grid(row=2, column=2,pady=5)


def upload():
    paymentlist = []
    adds = []
    global result1, pp, psy, dd, nd, lh, pyt, pya
    if lh == 1:
        sp1 = openpyxl.load_workbook(f'{nd}.xlsx')
    else:
        file = tk.filedialog.askopenfilename()
        sp1 = openpyxl.load_workbook(file)

    ofs = sp1["Sheet1"]
    ofs3 = sp1["Sheet2"]
    name1 = ofs.cell(3, 3)
    ofs.cell(2, 4, value=f"As of    { dt}")
    AS.cell(2, 3, value=f"As of  {dt}")
    n = name1.value
    ofs.delete_rows(idx=ofs.max_row)
    ofs.delete_rows(idx=ofs.max_row)
    ofs.delete_rows(idx=ofs.max_row)
    ofs.delete_rows(idx=ofs.max_row)
    ofs.delete_rows(idx=ofs.max_row)

    discount = tk.StringVar()
    discount1 = tk.StringVar()
    ofs.cell(2, 4, value=f"As of{dt}")
    status = ["Order", "Pullout"]
    my_w.iconify()

    window2 = tk.Toplevel(my_w)
    window2.geometry("675x510")
    window2.title(f"Order Form {n}")

    # treeview
    tv = ttk.Treeview(window2, selectmode='browse')
    tv.grid(row=7, column=1, columnspan=5, pady=10, padx=10)
    tv["columns"] = ("1", "2", "3", "4", "5", "6", "7")
    tv['show'] = 'headings'
    tv.column("1", width=50, anchor='c')
    tv.column("2", width=50, anchor='c')
    tv.column("3", width=50, anchor='c')
    tv.column("4", width=350, anchor='c')
    tv.column("5", width=50, anchor='c')
    tv.column("6", width=50, anchor='c')
    tv.column("7", width=50, anchor='c')
    tv.heading("1", text="DATE")
    tv.heading("2", text="DR/CM")
    tv.heading("3", text="QTY")
    tv.heading("4", text="Title of Books")
    tv.heading("5", text="Unit Price")
    tv.heading("6", text="Debit")
    tv.heading("7", text="Credit")
    click = tk.StringVar()
    click.set(status[0])
    e1_str = tk.StringVar()
    date = DateEntry(window2, selectmode="day")
    date.grid(row=1, column=4)
    label3 = tk.Label(window2, text="Tile of Books")
    label3.grid(row=0, column=2, columnspan=1)
    labeltotal = tk.Label(window2, text="Sub. Total :")
    labeltotal1 = tk.Label(window2, text="Less Discount :")
    labeltotal2 = tk.Label(window2, text="Total :")
    labeltotal2.grid(row=10, column=3, sticky="e")
    labeltotal1.grid(row=9, column=3, sticky="e")
    labeltotal.grid(row=8, column=3, sticky="e")
    total_discount = tk.Label(window2, fg="red", text=result1)
    total_discount.grid(row=9, column=4, sticky="e")
    total_total = tk.Label(window2, text=result1)
    total_total.grid(row=10, column=4, sticky="e")
    total_label = tk.Label(window2, text=result1)
    total_label.grid(row=8, column=4, sticky="e")
    label4 = tk.Label(window2, text="QTY")
    label4.grid(row=0, column=1)
    label5 = tk.Label(window2, text="DR#/CM#")
    label5.grid(row=0, column=3)
    label6 = tk.Label(window2, text="DATE")
    label6.grid(row=0, column=4)
    dr = tk.Entry(window2, width=6)

    dr.grid(row=1, column=3)
    e1 = tk.Entry(window2, textvariable=e1_str, width=55)
    e1.grid(row=1, column=2, columnspan=1)
    e2 = tk.Entry(window2, width=5)
    e2.grid(row=1, column=1)
    l1 = tk.Listbox(window2, height=6, relief="flat", bg="SystemButtonFace", width=55)
    l1.grid(row=2, column=2, columnspan=1, rowspan=4)
    dc = tk.Entry(window2, width=5, textvariable=discount)
    dc2 = tk.Entry(window2, width=5, textvariable=discount1)
    label8 = tk.Label(window2, text="    +")
    label8.grid(row=4, column=4, columnspan=5)
    dc1 = ofs3.cell(1,1)
    dcc = ofs3.cell(1,2)
    dc.insert(0, dc1.value)
    dc2.insert(0, dcc.value)
    dc2.grid(row=4, column=5)
    dc.grid(row=4, column=4)
    label7 = tk.Label(window2, text="Discount :")
    label7.grid(row=4, column=3)
    combo = tk.OptionMenu(window2, click, *status)
    combo.grid(row=3, column=3, padx=0, pady=0)

    for row1 in range(6, ofs.max_row):
        global num1

        cell = [ofs.cell(row1, 1), ofs.cell(row1, 2), ofs.cell(row1, 3), ofs.cell(row1, 4), ofs.cell(row1, 5), ofs.cell(row1, 6), ofs.cell(row1, 7)]

        if cell[6].value is None:
            cell[6].value = ""
            if cell[1].value == "ADD":
                dd = dd + int(cell[5].value)
                tv.insert("", 'end', iid=num1, values=(cell[0].value, cell[1].value, cell[2].value, cell[3].value, cell[4].value, cell[5].value, cell[6].value))
                num1 = num1 + 1
            else:
                tv.insert("", 'end', iid=num1, values=(cell[0].value, cell[1].value, cell[2].value, cell[3].value, cell[4].value, cell[5].value, cell[6].value))
                num1 = num1 + 1
                result1 = result1 + int(cell[5].value)


        if cell[5].value is None:
            cell[5].value = ""
            if cell[3].value == "Less Payment":
                psy = psy + int(cell[6].value)
                tv.insert("", 'end', iid=num1, values=(
                    cell[0].value, cell[1].value, cell[2].value, cell[3].value, cell[4].value, cell[5].value,
                    cell[6].value))
                num1 = num1 + 1

            else:
                if cell[0].value == "----":
                    num1 = num1 + 1
                else:
                    tv.insert("", 'end', iid=num1, values=(
                    cell[0].value, cell[1].value, cell[2].value, cell[3].value, cell[4].value, cell[5].value, cell[6].value))
                    result1 = result1 - int(cell[6].value)
                    num1 = num1 + 1

    def data_collect(self):
        e1.delete(0, END)
        e2.delete(0, END)
        dr.delete(0, END)
        date.delete(0, END)
        selected = tv.focus()

        values = tv.item(selected, 'values')
        e1.insert(0, values[3])
        e2.insert(0, values[2])
        dr.insert(0, values[1])
        date.insert(0, values[0])

    def up_button():
        global result1
        selected = tv.focus()
        values = tv.item(selected, 'values')
        my = e1.get()
        quantity = e2.get()
        y = 0
        if quantity == "":
            messagebox.showinfo("info", "Insert Quantity")

        if values[6] == "":
            result1 = result1 - int(values[5])
            for row in range(2, sheet.max_row + 1):
                cell1 = sheet.cell(row, 1)
                if my == cell1.value:
                    price = sheet.cell(row, 2)
                    qty1 = int(quantity)
                    total = qty1 * price.value
                    tv.item(selected, text="", values=(date.get(), dr.get(), quantity, cell1.value, price.value, total,""))
                    result1 += total
                    total_label.config(text="₱{:,.2f}".format(result1))
                    dis = int(dc.get())
                    dis2 = int(dc2.get())
                    dis1 = dis / 100 * result1
                    res = result1 - dis1
                    dis3 = dis2 / 100 * res
                    total_discount.config(text="₱{:,.2f}".format(dis1 + dis3))
                    res = res - dis3
                    total_total.config(text="₱{:,.2f}".format(res))
                else:
                    y += 1
                    if y == sheet.max_row - 1:
                        popup()

        else:
            result1 = int(values[6]) + result1
            for row in range(2, sheet.max_row + 1):
                cell1 = sheet.cell(row, 1)
                if my == cell1.value:
                    price = sheet.cell(row, 2)
                    qty1 = int(quantity)
                    total = qty1 * int(price.value)
                    tv.item(selected, text="", values=(date.get(), dr.get(), quantity, cell1.value, price.value, "",
                                                       total))
                    result1 -= total
                    total_label.config(text="₱{:,.2f}".format(result1))
                    dis = int(dc.get())
                    dis2 = int(dc2.get())
                    dis1 = dis / 100 * result1
                    res = result1 - dis1
                    dis3 = dis2 / 100 * res
                    total_discount.config(text="₱{:,.2f}".format(dis1 + dis3))
                    res = res - dis3
                    total_total.config(text="₱{:,.2f}".format(res))
                else:
                    y += 1
                    if y == sheet.max_row - 1:
                        popup()

    def my_upd(my_widget):
        my = my_widget.widget
        index = int(my.curselection()[0])
        value = my.get(index)
        e1_str.set(value)
        l1.delete(0, END)

    pp = result1
    pyt = pyt + psy
    pya = pya + dd


    def rem_button():
        global result1, pp, pyt, pya
        selected = tv.focus()
        values = tv.item(selected, 'values')

        if values[6] == "":
            if values[1] == "ADD":
                x = tv.selection()[0]
                tv.delete(x)
                total_label.config(text="₱{:,.2f}".format(result1))
                dis = int(dc.get())
                dis2 = int(dc2.get())
                dis1 = dis / 100 * result1
                res = result1 - dis1
                dis3 = dis2 / 100 * res
                total_discount.config(text="₱{:,.2f}".format(dis1 + dis3))
                pp = pp - int(values[5])
                pya = pya - int(values[5])
                total_total.config(text="₱{:,.2f}".format(pp))

            else:
                result1 = result1 - int(values[5])
                x = tv.selection()[0]
                tv.delete(x)
                total_label.config(text="₱{:,.2f}".format(result1))
                dis = int(dc.get())
                dis2 = int(dc2.get())
                dis1 = dis / 100 * result1
                pp = result1 - dis1
                dis3 = dis2 / 100 * pp
                pp = pp - dis3
                pp = pp - pyt
                pp = pp + pya
                total_discount.config(text="₱{:,.2f}".format(dis1 + dis3))
                total_total.config(text="₱{:,.2f}".format(pp))

        else:
            if values[3] == "Less Payment":
                x = tv.selection()[0]
                tv.delete(x)
                total_label.config(text="₱{:,.2f}".format(result1))
                dis = int(dc.get())
                dis2 = int(dc2.get())
                dis1 = dis / 100 * result1
                res = result1 - dis1
                dis3 = dis2 / 100 * res
                total_discount.config(text="₱{:,.2f}".format(dis1 + dis3))
                pp = pp + int(values[6])
                pyt = pyt - int(values[6])
                total_total.config(text="₱{:,.2f}".format(pp))

            else:
                result1 = int(values[6]) + result1
                x = tv.selection()[0]
                tv.delete(x)
                total_label.config(text="₱{:,.2f}".format(result1))
                dis = int(dc.get())
                dis2 = int(dc2.get())
                dis1 = dis / 100 * result1
                pp = result1 - dis1
                dis3 = dis2 / 100 * pp
                pp = pp - dis3
                pp = pp - pyt
                pp = pp + pya
                total_discount.config(text="₱{:,.2f}".format(dis1 + dis3))
                total_total.config(text="₱{:,.2f}".format(pp))
    def get_data(*args):
        search_str = e1.get()
        l1.delete(0, END)
        if search_str == "":
            l1.delete(0, END)

        else:
            for element in my_list:
                if re.match(search_str, element, re.IGNORECASE):
                    l1.insert(tk.END, element)

    def add_button(*args):
        tel = click.get()
        y = 0
        global num, result1, num1, pp
        if tel == "Order":
            my = e1.get()
            quantity = e2.get()
            dr1 = dr.get()
            dt1 = date.get()

            if quantity == "":
                messagebox.showinfo("info", "Insert Quantity")

            if dr1 == "":
                dr1 = "-"

            if dt1 == "":
                dt1 = "-"

            for row6 in range(2, sheet.max_row + 1):

                cell2 = sheet.cell(row6, 1)
                if my == cell2.value:
                    price = sheet.cell(row6, 2)
                    qty1 = int(quantity)
                    total = qty1 * price.value
                    tv.insert("", 'end', iid=num1,
                              values=(dt1, dr1, quantity, cell2.value, price.value, total, ""))
                    e1.delete(0, END)
                    date.delete(0, END)
                    dr.delete(0, END)

                    num1 = num1 + 1
                    result1 += total + 0.0
                    total_label.config(text="₱{:,.2f}".format(result1))
                    dis = int(dc.get())
                    dis2 = int(dc2.get())
                    dis1 = dis / 100 * result1
                    pp = result1 - dis1
                    dis3 = dis2 / 100 * pp
                    pp = pp - dis3
                    pp = pp - pyt
                    pp = pp + pya
                    total_discount.config(text="₱{:,.2f}".format(dis1 + dis3))
                    total_total.config(text="₱{:,.2f}".format(pp))

                else:
                    y += 1
                    if y == sheet.max_row - 1:
                        popup()

        if tel == "Pullout":

            my = e1.get()
            quantity = e2.get()
            dr1 = dr.get()
            dt1 = date.get()

            if quantity == "":
                messagebox.showinfo("info", "Insert Quantity")

            if dr1 == "":
                dr1 = "-"

            if dt1 == "":
                dt1 = "-"

            for row in range(2, sheet.max_row + 1):
                cell2 = sheet.cell(row, 1)
                if my == cell2.value:
                    price = sheet.cell(row, 2)
                    qty1 = int(quantity)
                    total = qty1 * int(price.value)
                    tv.insert("", 'end', iid=num1,
                              values=(dt1, dr1, quantity, cell2.value, price.value, "", total))
                    e1.delete(0, END)
                    date.delete(0, END)
                    dr.delete(0, END)
                    num1 = num1 + 1
                    result1 = result1 - total
                    total_label.config(text="₱{:,.2f}".format(result1))
                    dis = int(dc.get())
                    dis2 = int(dc2.get())
                    dis1 = dis / 100 * result1
                    pp = result1 - dis1
                    dis3 = dis2 / 100 * pp
                    pp = pp - dis3
                    pp = pp - pyt
                    pp = pp + pya
                    total_discount.config(text="₱{:,.2f}".format(dis1 + dis3))
                    total_total.config(text="₱{:,.2f}".format(pp))

                else:
                    y += 1
                    if y == sheet.max_row - 1:
                        popup()

    def sub_button():
        global result, result1, p, pt
        dis = int(dc.get())
        dis2 = int(dc2.get())

        for row2 in range(6, ofs.max_row+1):
            ofs.delete_rows(idx=6)

        for row4 in tv.get_children():
            values = tv.item(row4, 'values')
            ofs.append([values[0], values[1], values[2], values[3], values[4]])
            ss = ofs.max_row
            if values[3] == "Less Payment":
                paymentlist.append(row4)
                ofs.delete_rows(idx=ss)

            if values[1] == "ADD":
                adds.append(row4)
                ofs.delete_rows(idx=ss)

            if values[5] == "":
                tex = ofs.cell(int(ss), 4).value
                if tex is None:
                    ofs.delete_rows(idx=ss)

                else:
                    ofs.cell(ss, 7, value=int(values[6])).number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED2
                    nc = ofs.cell(int(ss), 7)
                    result = result - int(nc.value)
                    ofs.cell(ss, 8, value=result).number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED2
            else:
                tex1 = ofs.cell(int(ss), 4).value
                if tex1 is None:
                    ofs.delete_rows(idx=ss)
                else:
                    ofs.cell(ss, 6, value=int(values[5])).number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED2
                    nc = ofs.cell(int(ss), 6)
                    result = int(nc.value) + result
                    ofs.cell(ss, 8, value=result).number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED2

        dis1 = dis / 100 * result
        res = result - dis1
        dis3 = dis2 / 100 * res

        ofs.append(["----", "", "------", f"LESS {dc.get()}% DISCOUNT", "-----", "", "₱ - {:,.2f}".format(dis1),
                    "            {:,.2f}".format(res)])
        if int(dis2) == 0:

            pass

        else:
            res = res - dis3
            ofs.append(["----", "", "------", f"LESS Adtnl {dc2.get()}% DISCOUNT", "-----", "", "₱ - {:,.2f}".format(dis3),
                    "            {:,.2f}".format(res)])

        global jj
        jj = res
        for row5 in adds:
            mx = ofs.max_row + 1
            values = tv.item(row5, 'values')
            ofs.append([values[0], values[1], values[2], values[3], values[4], "", ""])
            ofs.cell(mx, 6, value=int(values[5])).number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED2
            nc = ofs.cell(int(mx), 6)
            jj = jj + int(nc.value)
            ofs.cell(mx, 8, value=jj).number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED2

        global ff
        ff = jj
        for row5 in paymentlist:

            mx = ofs.max_row + 1
            values = tv.item(row5, 'values')
            ofs.append([values[0], values[1], values[2], values[3], values[4], ""])
            ofs.cell(mx, 7, value=int(values[6])).number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED2
            nc = ofs.cell(int(mx), 7)
            ff = ff - int(nc.value)
            ofs.cell(mx, 8, value=ff).number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED2
        ofs.append(["----", "---", "---- ", "TOTAL BALANCE", "-----", "---", " --- ", "₱         {:,.2f}".format(ff)])
        ofs.append([""])
        ofs.append([""])
        ofs.append(["", "Received by:", "", "", "", "", "Prepared by:"])
        ofs.append([""])
        ofs.append(["", "____________", "", "", "", "", "_____________"])
        ofs3.cell(1, 1, value=dis)
        ofs3.cell(1, 2, value=dis2)
        AS.delete_rows(idx=AS.max_row)
        for row8 in range(5, AS.max_row+1):
            s = n
            a = row8 + 1
            cell3 = AS.cell(row8, 2)

            if s == cell3.value:
                a = a - 1

                AS.cell(a, 4, value=ff).number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED2

        for row9 in range(4, AS.max_row):
            x = row9 - 3
            cell9 = AS.cell(row9 + 1, 4)

            p = p + int(cell9.value)

            AS.cell(row9 + 1, 1, value=x)
            if row9 == AS.max_row - 1:
                AS.cell(row9 + 2, 3, value="TOTAL :")
                AS.cell(row9 + 2, 4, value=p).number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED2
        sp1.save(f'{n.upper()}.xlsx')
        sp2.save("sp.xlsx")
        if pt == 1:
            os.system(f'start "excel.exe" "{n.upper()}.xlsx"')
        else:
            messagebox.showinfo("info", "File Saved!!")

        my_w.destroy()

    def payment():
        window3 = tk.Toplevel(window2)
        window3.geometry("345x110")
        window3.title(f"Payment Form {n}")
        dts = DateEntry(window3, selectmode="day")
        dts.grid(row=3, column=1, pady=10, padx=10)
        pr = tk.Entry(window3, width=10)
        pr.grid(row=3, column=2, padx=10, pady=10)
        amt = tk.Entry(window3, width=20)
        amt.grid(row=3, column=3, padx=10, pady=10)
        lb1 = tk.Label(window3, text="DATE")
        lb2 = tk.Label(window3, text="PR#")
        lb3 = tk.Label(window3, text="AMOUNT")
        lb3.grid(row=2, column=3)
        lb2.grid(row=2, column=2)
        lb1.grid(row=2, column=1)
        global result1
        total_label.config(text="₱{:,.2f}".format(result1))

        def ok_button():
            global pyt
            amount = int(amt.get())
            pyt = amount + pyt
            global num, result1, num1, pp

            tv.insert("", 'end', iid=num1,
                      values=(dts.get(), pr.get(), "-", "Less Payment", "-", "", amount))
            num1 = num1 + 1
            pp = pp - amount

            total_total.config(text="₱{:,.2f}".format(pp))
            window3.destroy()

        oks = tk.Button(window3, text=" OK ", command=ok_button)
        oks.grid(row=4, column=2)

    def add_others():
        window3 = tk.Toplevel(window2)
        window3.geometry("345x300")
        window3.title(f"Payment Form {n}")
        dts = DateEntry(window3, selectmode="day")
        dts.grid(row=3, column=1, pady=10, padx=10)
        pr = tk.Entry(window3, width=20)
        pr.grid(row=3, column=2, padx=10, pady=10)
        amt = tk.Entry(window3, width=10)
        amt.grid(row=3, column=3, padx=10, pady=10)
        lb1 = tk.Label(window3, text="DATE")
        lb2 = tk.Label(window3, text="Details")
        lb3 = tk.Label(window3, text="AMOUNT")
        lb3.grid(row=2, column=3)
        lb2.grid(row=2, column=2)
        lb1.grid(row=2, column=1)

        def ok_button():
            global pya
            amount = int(amt.get())
            pya = pya + amount
            global num, result1, num1, pp
            tv.insert("", 'end', iid=num1,
                      values=(dts.get(),"ADD", "-", pr.get(), "-", amount,""))
            num1 = num1 + 1

            total_label.config(text="₱{:,.2f}".format(result1))
            pp = pp + amount
            total_total.config(text="₱{:,.2f}".format(pp))
            window3.destroy()

        oks = tk.Button(window3, text=" OK ", command=ok_button)
        oks.grid(row=4, column=2)

    def printcommand():
        global pt
        pt = 1
        sub_button()


    # Buttons
    b2 = tk.Button(window2, text="  Save ", command=sub_button)
    ps = tk.Button(window2, text="  Save and Print ", command=printcommand)
    b1 = tk.Button(window2, text="   Add   ", command=add_button)
    up = tk.Button(window2, text=" Update ", command=up_button)
    rem = tk.Button(window2, text="Remove", command=rem_button)
    pay = tk.Button(window2, text="Add Payment", command=payment)
    ad = tk.Button(window2, text="Add Others", command=add_others)

    rem.grid(row=8, column=1, columnspan=1)
    up.grid(row=6, column=3, columnspan=1)
    b2.grid(row=9, column=1, columnspan=2)
    ps.grid(row=9, column=2, columnspan=2)
    b1.grid(row=6, column=2, columnspan=1)
    pay.grid(row=8, column=1,columnspan=2, padx=10, pady=10)
    ad.grid(row=8, column=2, columnspan=3, padx=10, pady=10)
    l1.bind("<<ListboxSelect>>", my_upd)
    e1_str.trace('w', get_data)
    tv.bind("<<TreeviewSelect>>", data_collect)
    window2.bind("<Return>", add_button)

    def on_close():
        global lh, nd, result1, result, num, num3, num1, p, pymt, ff, jj, pp, psy, dd, pt
        num = 0
        num1 = 0
        num3 = 0
        result = 0
        result1 = 0
        pymt = 0
        ff = 0
        jj = 0
        p = 0
        pp = 0
        psy = 0
        dd = 0
        pt = 0
        window2.destroy()

    window2.protocol("WM_DELETE_WINDOW", on_close)


def view12():
    global lh, nd

    window4 = tk.Toplevel(my_w)
    window4.geometry("785x300")
    window4.title("All Schools")
    tv = ttk.Treeview(window4, selectmode='browse')
    tv.grid(row=2, column=1, columnspan=5, pady=10, padx=10)
    tv["columns"] = ("1", "2", "3", "4",)
    tv['show'] = 'headings'
    tv.column("1", width=30, anchor='c')
    tv.column("2", width=300, anchor='c')
    tv.column("3", width=350, anchor='c')
    tv.column("4", width=70, anchor='c')
    tv.heading("1", text="#")
    tv.heading("2", text="Name of School")
    tv.heading("3", text="Address")
    tv.heading("4", text="Balance")
    total = AS.cell(AS.max_row, 4)
    labeltotal2 = tk.Label(window4, text="Total :")
    labeltotal2.grid(row=10, column=3, sticky="e")
    total_total = tk.Label(window4, text="₱{:,.2f}".format(total.value))
    total_total.grid(row=10, column=4, sticky="e")

    for row1 in range(5, AS.max_row):
        global num3
        cell = [AS.cell(row1, 1), AS.cell(row1, 2), AS.cell(row1, 3), AS.cell(row1, 4)]
        tv.insert("", 'end', iid=num3, values=(cell[0].value, cell[1].value, cell[2].value, "₱{:,.2f}".format(cell[3].value)))
        num3 = num3 + 1

    def selectedfile():
        global lh, nd
        selected = tv.focus()
        values = tv.item(selected, 'values')
        lh = 1
        nd = values[1]
        upload()

    def printer():
        selected = tv.focus()
        values = tv.item(selected, 'values')
        nd1 = values[1]

        os.system(f'start "excel.exe" "{nd1.upper()}.xlsx"')

    def printer1():
        os.system(f'start "excel.exe" "sp.xlsx"')

    open2 = tk.Button(window4, text="  Open ", command=selectedfile)
    print1 = tk.Button(window4, text=" Print ", command=printer)
    print2 = tk.Button(window4, text=" Print Records", command=printer1)
    open2.grid(row=10, column=1, pady=10)
    print1.grid(row=10, column=2, pady=10)
    print2.grid(row=10, column=3, pady=10)

    def on_close():

        close = messagebox.askokcancel("Close", "Would you like to close the program?")
        if close:
            my_w.destroy()

    window4.protocol("WM_DELETE_WINDOW", on_close)


all1 = tk.Button(my_w, text=" View Schools ", command=view12)
recent = tk.Button(my_w, text=" Recent School ", command=upload)
okb = tk.Button(my_w, text="  OK  ", command=new_window)
okb.grid(row=3, column=2, pady=10)
recent.grid(row=5, column=2, sticky="w")
all1.grid(row=5, column=1)


my_w.mainloop()
